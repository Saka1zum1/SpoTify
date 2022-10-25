import json
import warnings
import requests
import spotipy
import sys
import openpyxl
import spotipy.util as util
from spotipy import SpotifyOAuth
import pandas as pd
import numpy as np
import seaborn as sns
from tqdm import tqdm
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn import metrics

SPOTIFY_USERNAME = ''  # 你的spotify账户名称
SPOTIFY_CLIENT_ID = ''  # 你在spotify developer创建的应用程序的用户id
SPOTIFY_CLIENT_SECRET = ''  # 你在spotify developer创建的应用程序的用户密码
SPOTIFY_REDIRECT_URI = 'http://localhost:8888/'  # 重定向网址，需要在应用程序设置里保持一致
SPOTIFY_SCOPE = 'user-library-read'  # 声明的app授权权限
SPOTIFY_USERID = ''  # 创建歌单时你的spotify用户id而不是用户名
SPOTIFY_SCOPE1 = 'playlist-modify-public,playlist-modify-private'  # 创建歌单的权限


class DataNotFoundException(Exception):
    pass


class CreateFailure(Exception):
    pass


class UpdateFailure(Exception):
    pass


class TokenError(Exception):
    pass


def _getToken(username, scope):
    auth_manager = SpotifyOAuth(client_id=SPOTIFY_CLIENT_ID,
                                client_secret=SPOTIFY_CLIENT_SECRET,
                                scope=scope,
                                username=username,
                                redirect_uri=SPOTIFY_REDIRECT_URI)

    sp = spotipy.Spotify(auth_manager=auth_manager)
    if sp:
        return sp
    else:
        message = 'Fail to get authorized(Perhaps the token is invalid )'
        raise TokenError(message)


class TrackInfo:
    def __init__(self, track_id=None, keyword=None, market=None, fetch=True):
        self.track_id = track_id
        self.keyword = keyword
        self.market = market
        if fetch:
            self._packData()

    def __repr__(self):
        if self.track_id or self.keyword:
            return " {}.{}(track_id={!r},keyword={!r})".format(
                self.__class__.__module__, self.__class__.__name__, self.track_id, self.keyword)

    def __str__(self):
        s = u" '%s' by %s\n%s\n%s" % (self.title, self.artist, self.genres, self.release_date)
        if sys.version_info.major < 3:
            return s.encode(getattr(sys.stdout, "encoding", "") or "utf8")
        else:
            return s

    def _getTrackData(self, sp, id, market):
        resp = sp.track(track_id=id, market=market)
        if resp:
            self.title = resp['name']
            self.album = resp['album']['name']
            self.image = resp['album']['images'][0]
            self.release_date = resp['album']['release_date']
            self.popularity = resp['popularity']
            self.markets = resp['available_markets']
            self.at_id = (resp['artists'][0]['uri']).split(':')[2]
            self.ab_id = (resp['album']['uri']).split(':')[2]
            if len(resp['artists']) != 1:
                self.artist = resp['artists'][0]['name']
                for at in resp['artists'][1:]:
                    self.artist += ', ' + at['name']
            else:
                self.artist = resp['artists'][0]['name']
        else:
            message = 'Track Not Found(Perhaps the track_id is invalid )'
            raise DataNotFound(message)

    def _getGenres(self, sp, id):
        resp = sp.artist(artist_id=id)
        if len(resp['genres']) != 0:
            self.genres = resp['genres'][0]
            if len(self.genres) != 1:
                for g in resp['genres'][1:]:
                    self.genres += '/' + g

            self.genres = self.genres.title()
        else:
            self.genres = 'Unclassified'

    def _getFeatures(self, sp, id):
        resp = sp.audio_features(tracks=id)
        self.features = {'danceability': resp[0]['danceability'], 'energy': resp[0]['energy'], 'key': resp[0]['key'],
                         'loudness': resp[0]['loudness'], 'speechiness': resp[0]['speechiness'],
                         'acousticness': resp[0]['acousticness'], 'instrumentalness': resp[0]['instrumentalness'],
                         'valence': resp[0]['valence'], 'tempo': resp[0]['tempo'], 'liveness': resp[0]['liveness'],
                         'length': str((resp[0]['duration_ms']) / 1000) + 's'}

    def _getAlbumData(self, sp, id):
        resp = sp.album(id)
        self.label = resp['label']
        cr = []
        for c in resp['copyrights']:
            cr.append(c['text'])
        cr = set(cr)
        for r in cr:
            self.copyrights = r

    def _searchTrack(self, sp, keyword):
        q = keyword.replace(' ', '%20')
        resp = sp.search(q, type='track', limit=1)
        self.track_id = (resp['tracks']['items'][0]['uri']).split(':')[2]

    def _packData(self):
        sp = _getToken(SPOTIFY_USERNAME, SPOTIFY_SCOPE)
        if self.keyword:
            self._searchTrack(sp, self.keyword)
        self._getTrackData(sp, self.track_id, market=self.market)
        self._getGenres(sp, self.at_id)
        self._getFeatures(sp, self.track_id)
        self._getAlbumData(sp, self.ab_id)


class AlbumInfo:
    def __init__(self, ab_id=None, keyword=None, fetch=True):
        self.ab_id = ab_id
        self.keyword = keyword

        if fetch:
            self._packData()

    def __repr__(self):
        if self.at_id or self.keyword:
            return " {}.{}(ab_id={!r},keyword={!r})".format(
                self.__class__.__module__, self.__class__.__name__, self.ab_id, self.keyword)

    def __str__(self):
        s = u"'%s' by %s\n%s\n%s\n%s\n%s" % (
            self.title, self.artist, self.genres, self.release_date, self.label, self.copyrights)
        if sys.version_info.major < 3:
            return s.encode(getattr(sys.stdout, "encoding", "") or "utf8")
        else:
            return s

    def _getAlbumData(self, sp, id):
        resp = sp.album(album_id=id)
        if resp is not None:
            self.title = resp['name']
            self.artist = resp['artists'][0]['name']
            self.genres = resp['genres']
            self.popularity = resp['popularity']
            self.label = resp['label']
            self.image = resp['images'][0]['url']
            self.release_date = resp['release_date']
            self.markets = resp['available_markets']
            self.at_id = resp['artists'][0]['uri'].split(':')[2]
            cr = []
            for c in resp['copyrights']:
                cr.append(c['text'])
            cr = set(cr)
            for r in cr:
                self.copyrights = r
            tracks = []
            for i in resp['tracks']['items']:
                track = {}
                id = i['uri'].split(':')[2]
                track[id] = i['name']
                tracks.append(track)
            self.tracks = tracks
        else:
            messages = 'Can not found the Album(Perhaps the ab_id is invalid?)'
            raise DataNotFoundException(messages)

    def _getGenres(self, sp, id):
        resp = sp.artist(artist_id=id)
        if len(resp['genres']) != 0:
            self.genres = resp['genres'][0]
            if len(self.genres) != 1:
                for g in resp['genres'][1:]:
                    self.genres += '/' + g

            self.genres = self.genres.title()
        else:
            self.genres = 'Unclassified'

    def _searchAlbum(self, sp, kw):
        q = kw.replace(' ', '%20')
        resp = sp.search(q, type='album', limit=1)
        if resp is not None:
            self.ab_id = resp['albums']['items'][0]['uri'].split(':')[2]
        else:
            messages = 'Album not found( Perhaps the name is misspelled? )'
            raise DataNotFoundException(messages)

    def _packData(self):
        sp = _getToken(SPOTIFY_USERNAME, SPOTIFY_SCOPE)
        if self.keyword is not None:
            self._searchAlbum(sp, self.keyword)

        self._getAlbumData(sp, self.ab_id)
        if len(self.genres) == 0:
            self._getGenres(sp, self.at_id)


class ArtistInfo:

    def __init__(self, at_id=None, keyword=None, fetch=True):
        self.at_id = at_id
        self.keyword = keyword

        if fetch:
            self._packData()

    def __repr__(self):
        if self.at_id or self.keyword:
            return " {}.{}(at_id={!r},keyword={!r})".format(
                self.__class__.__module__, self.__class__.__name__, self.at_id, self.keyword)

    def __str__(self):

        f = list(str(self.followers))
        f.insert(-3, ',')
        t = ''
        for ts in f:
            t += ts
        s = u"%s\n%s\n%s followers\n%d albums" \
            % (self.name, self.genres[0], t, len(self.albums))
        if sys.version_info.major < 3:
            return s.encode(getattr(sys.stdout, "encoding", "") or "utf8")
        else:
            return s

    def _getArtistData(self, sp, id):

        resp = sp.artist(artist_id=id)

        if resp is not None:
            self.name = resp['name']
            self.followers = resp['followers']['total']
            self.image = resp['images'][0]['url']
            self.popularity = resp['popularity']
            if len(resp['genres']) != 0:
                self.genres = resp['genres'][0]
                if len(self.genres) != 1:
                    for g in resp['genres'][1:]:
                        self.genres += '/' + g

                self.genres = self.genres.title()
            else:
                self.genres = 'Unclassified'
        else:
            messages = 'Can not found the artist (Perhaps the at_id is invalid )'
            raise DataNotFoundException(messages)

    def _getAlbumData(self, sp, id):

        resp = sp.artist_albums(artist_id=id)
        if resp is not None:

            albums = []

            for item in resp['items']:
                album = {}
                album_id = item['uri'].split(':')[2]
                album[album_id] = item['name']
                albums.append(album)

            self.albums = albums

        else:
            messages = 'Can not found the artist (Perhaps the at_id is invalid )'
            raise DataNotFoundException(messages)

    def _searchArtist(self, sp, kw):

        q = kw.replace(' ', '%20')
        resp = sp.search(q, type='artist', limit=1)
        if resp is not None:
            item = resp['artists']['items'][0]

            self.at_id = item['uri'].split(':')[2]
        else:
            messages = 'Artist not found( Perhaps the name is misspelled? )'
            raise DataNotFoundException(messages)

    def _packData(self):
        sp = _getToken(SPOTIFY_USERNAME, SPOTIFY_SCOPE)
        if self.keyword is not None:
            self._searchArtist(sp, self.keyword)

        self._getArtistData(sp, self.at_id)
        self._getAlbumData(sp, self.at_id)


class PlaylistInfo:
    def __init__(self, pl_id=None, keyword=None, fetch=True):
        self.pl_id = pl_id
        self.keyword = keyword

        if fetch:
            self._packData()

    def __repr__(self):
        if self.pl_id or self.keyword:
            return " {}.{}(pl_id={!r},keyword={!r})".format(
                self.__class__.__module__, self.__class__.__name__, self.pl_id, self.keyword)

    def __str__(self):
        if self.followers >= 1000:
            f = list(str(self.followers))
            f.insert(-3, ',')
            t = ''
            for ts in f:
                t += ts
            self.followers = t
        s = u"%s  (create by %s)\n%d tracks, %s followers\n%s" % (
            self.title, self.owner, self.total, self.followers, self.dsct)
        s += '-' * 80
        for track in self.tracks:
            for key in track.keys():
                s += "\n%d. '%s' by %s" % (self.tracks.index(track) + 1, key, track[key])
        if sys.version_info.major < 3:
            return s.encode(getattr(sys.stdout, "encoding", "") or "utf8")
        else:
            return s

    def __getitem__(self, index):
        return self.tr_id[index]

    def _getlist(self, sp, id):
        resp = sp.playlist(id)
        if resp:
            self.title = resp['name']
            self.followers = resp['followers']['total']
            self.total = resp['tracks']['total']
            self.dsct = resp['description']
            self.owner = resp['owner']['display_name']
            self.image = resp['images'][0]['url']
            self.tracks = []
            self.tr_id = []
            for item in resp['tracks']['items']:
                track = item['track']['name']
                self.tr_id.append(item['track']['uri'].split(':')[2])
                if len(item['track']['artists']) != 1:
                    artist = item['track']['artists'][0]['name']
                    for at in item['track']['artists'][1:]:
                        artist += ', ' + at['name']
                else:
                    artist = item['track']['artists'][0]['name']

                self.tracks.append({track: artist})

    def _searchlist(self, sp, kw):
        q = kw.replace(' ', '%20')
        resp = sp.search(q, type='playlist', limit=1)
        self.pl_id = resp['playlists']['items'][0]['uri'].split(':')[2]

    def _packData(self):
        sp = _getToken(SPOTIFY_USERNAME, SPOTIFY_SCOPE)
        if self.keyword is not None:
            self._searchlist(sp, self.keyword)
        self._getlist(sp, self.pl_id)


def _createPlaylist(name, description):
    sp = _getToken(SPOTIFY_USERID, SPOTIFY_SCOPE1)
    try:
        resp = sp.user_playlist_create(SPOTIFY_USERID, name, public=True, collaborative=False, description=description)
        print('You have created a new playlist!')
        playlist_id = resp['uri'].split(':')[2]
        return playlist_id

    except:
        messages = 'Failed to create playlist(Perhaps the USERID is invalid)'
        raise CreateFailure(messages)


def _add_items_to_Playlist(pl_id, items):
    sp = _getToken(SPOTIFY_USERID, SPOTIFY_SCOPE1)
    try:
        sp.playlist_add_items(playlist_id=pl_id, items=items)
    except:
        messages = 'Failed to add items in the playlist(Perhaps the pl_id is invalid?)'
        raise UpdateFailure(messages)
    finally:
        print('Update finished')


def _get_track_uri():
    sp = _getToken(SPOTIFY_USERNAME, SPOTIFY_SCOPE)
    wb = openpyxl.load_workbook('decade-charts.xlsx')

    ws = wb['10s']
    for raw in range(2, 102):
        keyword = ws['B' + str(raw)].value.lower() + ' artist:' + ws['D' + str(raw)].value
        if keyword is not None:
            try:
                uri = sp.search(keyword, type='track', limit=1)['tracks']['items'][0]['uri'].split(':')[2]
                ws['K' + str(raw)].value = uri
            except:
                continue

    wb.save('decade-charts.xlsx')


def _getitems(sheet):
    wb = openpyxl.load_workbook('decade-charts.xlsx')
    ws = wb[sheet]
    uris = []
    for i in range(2, 102):
        uris.append('spotify:track:%s' % ws['K' + str(i)].value)
    return uris


def set_color(n):
    colors = []
    for i in range(n):
        i = lambda: random.randint(0, 255)
        color = '#%02X%02X%02X' % (i(), i(), i())
        colors.append(color)
    return colors
